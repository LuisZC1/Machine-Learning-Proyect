# -*- coding: utf-8 -*-
"""
Created on Mon Jul 24 14:15:24 2023

@author: luisz
"""

from openpyxl import load_workbook

# Cargar el archivo de Excel
wb = load_workbook(filename='EC_paro_2022_hate_speech.xlsx')

# Seleccionar la hoja en la que trabajarás. Si no sabes el nombre, puedes usar wb.sheetnames para obtener una lista de todas las hojas.
sheet = wb['Sample_hate_speech']

# Definir las variables que contarán los diferentes casos
verdadero_positivo = 0
falso_positivo = 0
verdadero_negativo = 0
falso_negativo = 0
total_positivos_J = 0  # Nuevo contador para el total de positivos en J

# Recorrer cada fila en la hoja (a partir de la segunda fila, asumiendo que la primera fila contiene encabezados)
for i in range(2, sheet.max_row+1):
    celda_I = sheet.cell(row=i, column=9)
    celda_J = sheet.cell(row=i, column=10)

    # Comprobar el valor en cada celda y actualizar las variables correspondientes
    if celda_I.value == 1 and celda_J.value == 1:
        verdadero_positivo += 1
    elif celda_I.value == 0 and celda_J.value == 1:
        falso_positivo += 1
    elif celda_I.value == 1 and celda_J.value == 0:
        falso_negativo += 1
    elif celda_I.value == 0 and celda_J.value == 0:
        verdadero_negativo += 1

    # Contar el total de positivos en J
    if celda_J.value == 1:
        total_positivos_J += 1

# Calcular el porcentaje de verdaderos y falsos positivos en relación al total de positivos en J
porcentaje_verdaderos_positivos = (verdadero_positivo / total_positivos_J) * 100 if total_positivos_J > 0 else 0
porcentaje_falsos_positivos = (falso_positivo / total_positivos_J) * 100 if total_positivos_J > 0 else 0

# Imprimir los resultados
print(f"Verdaderos positivos: {verdadero_positivo}")
print(f"Falsos positivos: {falso_positivo}")
print(f"Falsos negativos: {falso_negativo}")
print(f"Verdaderos negativos: {verdadero_negativo}")
print(f"Total de positivos en J: {total_positivos_J}")
print(f"Porcentaje de verdaderos positivos: {porcentaje_verdaderos_positivos}%")
print(f"Porcentaje de falsos positivos: {porcentaje_falsos_positivos}%")

