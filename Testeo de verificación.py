# -*- coding: utf-8 -*-
"""
Created on Mon Jul 24 13:25:30 2023

@author: luisz
"""

from openpyxl import load_workbook
import re

# Cargar el archivo de Excel
wb = load_workbook(filename='EC_paro_2022_hate_speech.xlsx')

# Seleccionar la hoja en la que trabajarás. Si no sabes el nombre, puedes usar wb.sheetnames para obtener una lista de todas las hojas.
sheet = wb['Sample_hate_speech']

# Definir las palabras o frases a buscar
palabras = [
    'Izapreso',
    'Izavandaliza',
    'Izaterrorista',
    'IzaGolpista',
    'Leonidas Iza por terrorismo',
    'Iza le mintió',
    'Iza debe pagar',
    'terrorista de Iza',
    ' indios salvajes',
    'Ya no mas Iza',
    'es culpa de Iza',
    'Maldito Iza',
    'Iza es un traidor',
    'Iza y la CONAIE han  engañado al país',
    'Iza ha engañado al país',
    'CONAIE ha engañado al país',
    'INDIOS ASESINOS TERRORISTAS',
    'INDIOS TERRORISTAS',
    'INDIOS ASESINOS',
    'ConaieTerrorista',
    'Izaasesino',
    'IzaVandaliza',
    'Iza engaña al pueblo',
    'caprichos de leonidas Iza',
    'Iza es un líder mentiroso',
    'Iza es un mentiroso',
    ' terrorismo de Iza y la Conaie',
    ' terrorismo de Iza',
    ' terrorismo de la Conaie',
    ' Iza siguen destruyendo el país',
    ' Iza sigue destruyendo el país',
    'ViolenciaConaie',
    'IZA NO TIENE INTERÉS EN EL PUEBLO',
    'Iza traidor, mentiroso',
    'Iza traidor',
    'Iza mentiroso',
    'Indio de mierda',
    ' Iza es un miserable y vil mentiroso',
    ' Iza es un miserable',
    ' Iza es un vil mentiroso',
    ' Iza es un mentiroso',
    'LeonidasIzaSal1 traicionero',
    ' Iza es un farsante',
    'NarcoGolpe',
    'TodoCorreistaEsLadron',
    ' Iza traicionó a su pueblo',
    'engañar por leonidas Iza',
    'IzaCojudo',
    'FueraIzaFuera',
    'bobos indígenas',
    'Iza no está dispuesto a dialogar',
    'COMUNISMO INDOAMERICANO'
]


# Para cada palabra o frase, si se encuentra en la columna B (ignorando mayúsculas, minúsculas y caracteres especiales), poner 1 en la columna J
# Si no se encuentra ninguna palabra, poner 0 en la columna J
for i in range(2, sheet.max_row+1):  # Empezamos desde 2 asumiendo que hay encabezados en la primera fila
    celda_B = sheet.cell(row=i, column=2)
    texto_B = re.sub(r'\W+', ' ', str(celda_B.value)).lower() if celda_B.value is not None else ""
    sheet.cell(row=i, column=10, value=0)  # Ahora la columna J se inicializa con '0'
    for palabra in palabras:
        palabra = re.sub(r'\W+', ' ', palabra).lower()
        if palabra in texto_B:
            sheet.cell(row=i, column=10, value=1)  # Si se encuentra una palabra, se cambia a '1'
            break  # Si ya encontramos una palabra, no necesitamos buscar las otras

# Guardar los cambios en el mismo archivo de Excel
wb.save('EC_paro_2022_hate_speech.xlsx')
